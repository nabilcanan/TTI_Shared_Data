import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# DO we use KIMBALL PN_MANUFACTURER_MFPN or KIMBALL PN


#  Kimball
def kimball_logic():
    # Create a new Toplevel window for Benchmark
    kimball_window = tk.Toplevel()
    kimball_window.title("Kimball Module")
    kimball_window.geometry("800x500")

    def open_and_process_file():
        # Read the first file containing CPNs
        first_file_path = filedialog.askopenfilename(title="Select CPN File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if first_file_path:
            df_first = pd.read_excel(first_file_path)  # Load the file into a DataFrame
            if 'CPN' in df_first.columns:
                # Remove duplicates based on the 'CPN' column, this is column A, but we search by header title
                df_first = df_first.drop_duplicates(subset=['CPN'])
                # Continue processing with other files if the first file is processed successfully
                if process_second_file(df_first) and process_third_file(df_first) and process_fourth_file(df_first):
                    # Save the final processed DataFrame
                    save_file(df_first)
            else:
                # Error message if 'CPN' column is not found
                messagebox.showerror("Error", "CPN column not found in the first file.")
        else:
            # Warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "First file open cancelled.")

    def process_second_file(df_first):
        # Not actual Kimball contract, we didn't have one, So we used the creation program to finish this module
        # Open a dialog to select the second file, typically the latest contract file for Kimball
        second_file_path = filedialog.askopenfilename(title="Select your LATEST Contract File for Kimball",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if second_file_path:
            # Load the file into a DataFrame without setting a header initially
            df_second_raw = pd.read_excel(second_file_path, header=None)

            # Find the row index where the desired header ('Kimball part #') is located
            header_row_idx = None
            for idx, row in df_second_raw.iterrows():
                if 'IPN' in row.values:
                    header_row_idx = idx
                    break

            if header_row_idx is not None and isinstance(header_row_idx, int):
                # Re-read the file with the correct header row
                df_second = pd.read_excel(second_file_path, header=int(header_row_idx))
                # Clean and process data
                df_first['CPN'] = df_first['CPN'].astype(str).str.strip()
                df_second['IPN'] = df_second['IPN'].astype(str).str.strip()

                # Proceed with processing if 'IPN' column is found
                if 'IPN' in df_second.columns:
                    df_first['On Contract'] = df_first['CPN'].isin(df_second['IPN']).map({True: 'Y', False: ''})
                    return True
                else:
                    messagebox.showerror("Error", "IPN column not found in the second file.")
                    return False
            else:
                messagebox.showerror("Error", "Header row with 'IPN' not found in the file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Second file open cancelled.")
            return False

    def process_third_file(df_first):
        # Open a dialog to select the third file, typically the Kimball Backlog File
        third_file_path = filedialog.askopenfilename(title="Select your Kimball Backlog File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if third_file_path:
            # Load the third file into a DataFrame
            df_third = pd.read_excel(third_file_path)
            # Check if the required columns are in the third file
            if 'Backlog CPN' in df_third.columns and 'Scheduled Arrival' in df_third.columns:
                # Group the data by 'Backlog CPN' and get the latest (max) 'Scheduled Arrival' for each CPN
                latest_arrivals = df_third.groupby('Backlog CPN')['Scheduled Arrival'].max()
                # Map the latest arrival dates to the corresponding CPNs in the first DataFrame
                df_first['Last Ship Date'] = df_first['CPN'].map(latest_arrivals)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the third file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Third file open cancelled.")
            return False

    def process_fourth_file(df_first):
        # Open a dialog to select the fourth file, typically the Kimball Sales History File
        fourth_file_path = filedialog.askopenfilename(title="Select your Kimball Sales History File",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if fourth_file_path:
            # Load the fourth file into a DataFrame
            df_fourth = pd.read_excel(fourth_file_path)
            # Check if the required columns are in the fourth file
            if 'Last Ship CPN' in df_fourth.columns and 'Last Ship Date' in df_fourth.columns:
                # Group the data by 'Last Ship CPN' and get the latest (max) 'Last Ship Date' for each CPN
                latest_ship_dates = df_fourth.groupby('Last Ship CPN')['Last Ship Date'].max()
                # Update 'Last Ship Date' in the first DataFrame with the latest ship date where it is missing
                df_first['Last Ship Date'] = df_first.apply(
                    lambda row: row['Last Ship Date'] if pd.notna(row['Last Ship Date']) else latest_ship_dates.get(
                        row['CPN']), axis=1)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the fourth file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Fourth file open cancelled.")
            return False

    def save_file(df_first):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_file_path:
            # Save DataFrame to an Excel file
            df_first.to_excel(save_file_path, index=False)

            # Open the saved Excel file and adjust column widths
            workbook = load_workbook(save_file_path)
            worksheet = workbook.active

            # Auto-adjust the width of Column A (or any other column as needed)
            column_widths = []
            for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=worksheet.max_row):
                for cell in row:
                    try:
                        # Get the length of the cell value
                        column_widths.append(len(str(cell.value)))
                    except StopIteration:  # StopIteration should work here to raise an error
                        "You messed up"
                        pass

            # Set the width of Column A to the width of the longest entry
            if column_widths:
                max_width = max(column_widths)
                worksheet.column_dimensions[get_column_letter(1)].width = max_width

            # Save the changes
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "Files processed, formatted, and saved successfully!")
        else:
            messagebox.showwarning("Cancelled", "File save cancelled.")

    # Modern color scheme
    bg_color = '#3b5998'  # A shade of blue
    text_color = '#ffffff'  # White for readability
    button_color = '#8b9dc3'  # Lighter shade of blue for the button
    button_text_color = '#ffffff'  # White text on the button

    # Apply background color
    kimball_window.configure(bg=bg_color)

    # Add content specific to the Benchmark module you turned her against me,
    # you have done that yourself ! You've allowed this dark lord to twist you mind.
    label = tk.Label(kimball_window, text="Welcome Partnership Member", font=("Verdana", 24), bg=bg_color,
                     fg=text_color)
    label.pack(pady=20)

    # Instructions for program
    instructions = ("Instructions:\n"
                    "1. Select the file containing the CPNs.\n"
                    "2. Choose the Latest Contract File for Kimball.\n"
                    "3. Select the Kimball Backlog File.\n"
                    "4. Pick the Kimball Sales History File.\n"
                    "5. Remember to SAVE your final file upon completion.")

    label = tk.Label(kimball_window, text=instructions, font=("Verdana", 20), bg=bg_color, fg=text_color)
    label.pack(pady=20)

    # Button to open the file dialog
    open_file_btn = tk.Button(kimball_window, text="Select Excel Files", command=open_and_process_file,
                              bg=button_color, fg=button_text_color, font=("Verdana", 16))
    open_file_btn.pack(pady=10)
