import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def press_action():
    print("hello partnership")
    pass


# Sanmina
# Completed

def sanmina_logic():
    # Create a new Toplevel window for Benchmark
    sanmina_window = tk.Toplevel()
    sanmina_window.title("Sanmina Module")
    sanmina_window.geometry("800x500")

    def open_and_process_file():
        # Read the first file containing CPNs
        first_file_path = filedialog.askopenfilename(title="Select CPN File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if first_file_path:
            df_first = pd.read_excel(first_file_path)  # Load the file into a DataFrame
            if 'CPN' in df_first.columns:
                # Remove duplicates based on the 'CPN' column
                df_first = df_first.drop_duplicates(subset=['CPN'])

                # Continue processing with other files if the first file is processed successfully
                if process_second_file(df_first):
                    if process_third_file(df_first):
                        if process_fourth_file(df_first):
                            if process_agility_file(df_first):
                                # Save the final processed DataFrame
                                save_file(df_first)
                            else:
                                messagebox.showwarning("Warning", "Agility file processing was not completed.")
                        else:
                            messagebox.showwarning("Warning", "Fourth file processing was not completed.")
                    else:
                        messagebox.showwarning("Warning", "Third file processing was not completed.")
                else:
                    messagebox.showwarning("Warning", "Second file processing was not completed.")
            else:
                # Error message if 'CPN' column is not found
                messagebox.showerror("Error", "CPN column not found in the first file.")
        else:
            # Warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "First file open cancelled.")

    def process_agility_file(df_first):
        # Open a dialog to select the Agility file
        agility_file_path = filedialog.askopenfilename(
            title="Select your Agility File For Sanmina",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )

        if agility_file_path:
            # Load the Agility file into a DataFrame
            df_agility = pd.read_excel(agility_file_path)

            # Check if 'Item Number' column exists in the Agility file
            if 'Item Number' in df_agility.columns:
                # Update 'On Contract' column in df_first based on 'Item Number' in the Agility file
                df_first['On Contract'] = df_first.apply(
                    lambda row: 'Y' if row['CPN'] in df_agility['Item Number'].values else row['On Contract'], axis=1)
                return True
            else:
                messagebox.showerror("Error", "Item Number column not found in the Agility file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Agility file selection cancelled.")
            return False

    def process_second_file(df_first):
        second_file_path = filedialog.askopenfilename(title="Select your LATEST AWARD File for Sanmina",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if second_file_path:
            df_second_raw = pd.read_excel(second_file_path, header=None)

            # Attempt to find the header row
            header_row_idx = None
            for idx, row in df_second_raw.iterrows():
                if 'Item Number' in row.values:
                    header_row_idx = idx
                    break

            if header_row_idx is not None and isinstance(header_row_idx, int):
                df_second = pd.read_excel(second_file_path, header=int(header_row_idx))
                # Clean and process data
                df_first['CPN'] = df_first['CPN'].astype(str).str.strip()
                df_second['Item Number'] = df_second['Item Number'].astype(str).str.strip()

                # Update 'On Contract' column
                df_first['On Contract'] = df_first['CPN'].isin(df_second['Item Number']).map({True: 'Y', False: ''})
                return True
            else:
                messagebox.showerror("Error", "Header row with 'Item Number' not found in the file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Second file open cancelled.")
            return False

    def process_third_file(df_first):
        # Open a dialog to select the third file, typically the Sanmina Backlog File
        third_file_path = filedialog.askopenfilename(title="Select your Sanmina Backlog File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if third_file_path:
            # Load the third file into a DataFrame
            df_third = pd.read_excel(third_file_path)
            # Check if the required columns are in the third file
            if 'Backlog CPN' in df_third.columns and 'Scheduled Arrival' in df_third.columns:
                # Group the data by 'Backlog CPN' and get the latest (max) 'Scheduled Arrival' for each CPN
                latest_arrivals = df_third.groupby('Backlog CPN')['Scheduled Arrival'].max()
                # Map the latest arrival dates to the corresponding CPNs in the first DataFrame
                df_first['Last Ship Date'] = df_first['CPN'].map(latest_arrivals)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the third file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Third file open cancelled.")
            return False

    def process_fourth_file(df_first):
        # Open a dialog to select the fourth file, typically the Sanmina Sales History File
        fourth_file_path = filedialog.askopenfilename(title="Select your Sanmina Sales History File",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if fourth_file_path:
            # Load the fourth file into a DataFrame
            df_fourth = pd.read_excel(fourth_file_path)
            # Check if the required columns are in the fourth file
            if 'Last Ship CPN' in df_fourth.columns and 'Last Ship Date' in df_fourth.columns:
                # Group the data by 'Last Ship CPN' and get the latest (max) 'Last Ship Date' for each CPN
                latest_ship_dates = df_fourth.groupby('Last Ship CPN')['Last Ship Date'].max()
                # Update 'Last Ship Date' in the first DataFrame with the latest ship date where it is missing
                df_first['Last Ship Date'] = df_first.apply(
                    lambda row: row['Last Ship Date'] if pd.notna(row['Last Ship Date'])
                    else latest_ship_dates.get(row['CPN']), axis=1)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the fourth file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Fourth file open cancelled.")
            return False

    def save_file(df_first):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_file_path:
            # Save DataFrame to an Excel file
            df_first.to_excel(save_file_path, index=False)

            # Open the saved Excel file and adjust column widths
            workbook = load_workbook(save_file_path)
            worksheet = workbook.active

            # Auto-adjust the width of Column A (or any other column as needed)
            column_widths = []
            for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=worksheet.max_row):
                for cell in row:
                    try:
                        # Get the length of the cell value
                        column_widths.append(len(str(cell.value)))
                    except StopIteration:  # StopIteration should work here to raise an error
                        "You messed up"
                        pass

            # Set the width of Column A to the width of the longest entry
            if column_widths:
                max_width = max(column_widths)
                worksheet.column_dimensions[get_column_letter(1)].width = max_width

            # Save the changes
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "Files processed, formatted, and saved successfully!")
        else:
            messagebox.showwarning("Cancelled", "File save cancelled.")

    # Modern color scheme
    bg_color = '#3b5998'  # A shade of blue
    text_color = '#ffffff'  # White for readability
    button_color = '#8b9dc3'  # Lighter shade of blue for the button
    button_text_color = '#ffffff'  # White text on the button

    # Apply background color
    sanmina_window.configure(bg=bg_color)

    # Add content specific to the Benchmark module
    label = tk.Label(sanmina_window, text="Welcome Partnership Member", font=("Verdana", 24), bg=bg_color,
                     fg=text_color)
    label.pack(pady=20)

    # Instructions for program
    instructions = ("Instructions:\n"
                    "1. Select the file containing the CPNs.\n"
                    "2. Choose the Latest Contract File for Sanmina.\n"
                    "3. Select the Sanmina Backlog File.\n"
                    "4. Pick the Sanmina Sales History File.\n"
                    "5. Finally Select you Sanmina Agility File\n"
                    "5. Remember to SAVE your final file upon completion.")

    label = tk.Label(sanmina_window, text=instructions, font=("Verdana", 20), bg=bg_color, fg=text_color)
    label.pack(pady=20)

    # Button to open the file dialog
    open_file_btn = tk.Button(sanmina_window, text="Select Excel Files", command=open_and_process_file,
                              bg=button_color, fg=button_text_color, font=("Verdana", 16))
    open_file_btn.pack(pady=10)

    press_action()
