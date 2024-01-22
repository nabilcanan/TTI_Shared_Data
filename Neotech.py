import tkinter as tk
from tkinter import filedialog, messagebox, font
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def neotech_logic():
    # Create a new Toplevel window
    neotech_window = tk.Toplevel()
    neotech_window.title("Neotech Module")
    neotech_window.geometry("800x500")

    def open_and_process_file():
        # Read the first file
        first_file_path = filedialog.askopenfilename(title="Select CPN File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if first_file_path:
            df_first = pd.read_excel(first_file_path)
            if 'CPN' in df_first.columns:
                df_first = df_first.drop_duplicates(subset=['CPN'])

                # Process the second, third, and fourth files
                if process_second_file(df_first) and process_third_file(df_first) and process_fourth_file(df_first):
                    # Save the processed file
                    save_file(df_first)
            else:
                messagebox.showerror("Error", "CPN column not found in the first file.")
        else:
            messagebox.showwarning("Cancelled", "First file open cancelled.")

    def process_second_file(df_first):
        second_file_path = filedialog.askopenfilename(title="Select your LATEST Contract File for Neotech",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if second_file_path:
            df_second = pd.read_excel(second_file_path)
            if 'PartNum' in df_second.columns:
                df_first['On Contract'] = df_first['CPN'].isin(df_second['PartNum']).map({True: 'Y', False: ''})
                return True
            else:
                messagebox.showerror("Error", "PartNum column not found in the second file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Second file open cancelled.")
            return False

    def process_third_file(df_first):
        third_file_path = filedialog.askopenfilename(title="Select your Neotech Backlog File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if third_file_path:
            df_third = pd.read_excel(third_file_path)
            if 'Backlog CPN' in df_third.columns and 'Scheduled Arrival' in df_third.columns:
                latest_arrivals = df_third.groupby('Backlog CPN')['Scheduled Arrival'].max()
                df_first['Last Ship Date'] = df_first['CPN'].map(latest_arrivals)
                return True
            else:
                messagebox.showerror("Error", "Required columns not found in the third file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Third file open cancelled.")
            return False

    def process_fourth_file(df_first):
        fourth_file_path = filedialog.askopenfilename(title="Select your Neotech Sales History File",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if fourth_file_path:
            df_fourth = pd.read_excel(fourth_file_path)
            if 'Last Ship CPN' in df_fourth.columns and 'Last Ship Date' in df_fourth.columns:
                # Change here: using max() to get the latest 'Last Ship Date'
                latest_ship_dates = df_fourth.groupby('Last Ship CPN')['Last Ship Date'].max()

                # Update 'Last Ship Date' with the latest 'Last Ship Date' where 'Last Ship Date' is missing
                df_first['Last Ship Date'] = df_first.apply(
                    lambda row: row['Last Ship Date'] if pd.notna(row['Last Ship Date']) else latest_ship_dates.get(
                        row['CPN']), axis=1)
                return True
            else:
                messagebox.showerror("Error", "Required columns not found in the fourth file.")
                return False
        else:
            messagebox.showwarning("Cancelled", "Fourth file open cancelled.")
            return False

    def save_file(df_first):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_file_path:
            # Save DataFrame to an Excel file
            df_first.to_excel(save_file_path, index=False)

            # Open the saved Excel file and add formatting
            workbook = load_workbook(save_file_path)
            worksheet = workbook.active

            # Apply wrap text for specific columns just going to do A for now and implement more later
            for row in worksheet.iter_rows(min_row=2):  # Skip the header row
                for cell in row:
                    if cell.column_letter in ['A']:
                        cell.alignment = Alignment(wrap_text=False) # Right now we are set to false considering it's not working properly

            # Enable filters for all columns
            worksheet.auto_filter.ref = worksheet.dimensions

            # Save the changes
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "Files processed, formatted, and saved successfully!")
        else:
            messagebox.showwarning("Cancelled", "File save cancelled.")

    # Add content specific to the Benchmark module
    label = tk.Label(neotech_window, text="Welcome Partnership Member", font=("Verdana", 24))
    label.pack(pady=20)

    # Add content specific to the Benchmark module
    my_font = font.Font(family="Verdana", size=20, underline=True)  # Define a font with underline
    label = tk.Label(neotech_window, text="Instructions:", font=my_font)
    label.pack(pady=20)

    # Add content specific to the Benchmark module
    label = tk.Label(neotech_window, text="First you will select the file where we have our CPN's\n"
                                          "Next we will select the Latest Contract File for Neotech\n"
                                          "Next we will select our Neotech Backlog File\n"
                                          "Lastly we will select out Neotech Sales History File\n"
                                          "Finally SAVE your final file", font=("Verdana", 20))
    label.pack(pady=20)

    # Button to open the file dialog
    open_file_btn = tk.Button(neotech_window, text="Select Excel Files", command=open_and_process_file)
    open_file_btn.pack(pady=10)
