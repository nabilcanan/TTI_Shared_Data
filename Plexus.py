import tkinter as tk
from tkinter import filedialog, messagebox, font
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


#  Plexus
def plexus_logic():
    # Create a new Toplevel window for Benchmark
    plexus_window = tk.Toplevel()
    plexus_window.title("Plexus Module")
    plexus_window.geometry("800x500")

    # Add content specific to the Benchmark module
    label = tk.Label(plexus_window, text="This is the Plexus module", font=("Verdana", 16))
    label.pack(pady=20)

    def open_and_process_file():
        # Read the first file containing CPNs
        first_file_path = filedialog.askopenfilename(title="Select CPN File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if first_file_path:
            df_first = pd.read_excel(first_file_path)  # Load the file into a DataFrame
            if 'CPN' in df_first.columns:
                # Remove duplicates based on the 'CPN' column
                df_first = df_first.drop_duplicates(subset=['CPN'])
                # Continue processing with other files if the first file is processed successfully
                if process_second_file(df_first) and process_third_file(df_first) and process_fourth_file(df_first):
                    # Save the final processed DataFrame
                    save_file(df_first)
            else:
                # Error message if 'CPN' column is not found
                messagebox.showerror("Error", "CPN column not found in the first file.")
        else:
            # Warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "First file open cancelled.")

    def process_second_file(df_first):
        # Read the second file, typically the latest contract file
        second_file_path = filedialog.askopenfilename(title="Select your LATEST Contract File for Plexus",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if second_file_path:
            df_second = pd.read_excel(second_file_path)  # Load the file into a DataFrame
            if 'PartNum' in df_second.columns:
                # Check if CPNs from the first file are present in the 'PartNum' column of the second file
                # Mark 'Y' in 'On Contract' column if present, otherwise leave blank
                df_first['On Contract'] = df_first['CPN'].isin(df_second['PartNum']).map({True: 'Y', False: ''})
                return True
            else:
                # Error message if 'PartNum' column is not found
                messagebox.showerror("Error", "PartNum column not found in the second file.")
                return False
        else:
            # Warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Second file open cancelled.")
            return False

    def process_third_file(df_first):
        # Open a dialog to select the third file, typically the Plexus Backlog File
        third_file_path = filedialog.askopenfilename(title="Select your Plexus Backlog File",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
        if third_file_path:
            # Load the third file into a DataFrame
            df_third = pd.read_excel(third_file_path)
            # Check if the required columns are in the third file
            if 'Backlog CPN' in df_third.columns and 'Scheduled Arrival' in df_third.columns:
                # Group the data by 'Backlog CPN' and get the latest (max) 'Scheduled Arrival' for each CPN
                latest_arrivals = df_third.groupby('Backlog CPN')['Scheduled Arrival'].max()
                # Map the latest arrival dates to the corresponding CPNs in the first DataFrame
                df_first['Last Ship Date'] = df_first['CPN'].map(latest_arrivals)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the third file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Third file open cancelled.")
            return False

    def process_fourth_file(df_first):
        # Open a dialog to select the fourth file, typically the Plexus Sales History File
        fourth_file_path = filedialog.askopenfilename(title="Select your Plexus Sales History File",
                                                      filetypes=[("Excel files", "*.xlsx;*.xls")])
        if fourth_file_path:
            # Load the fourth file into a DataFrame
            df_fourth = pd.read_excel(fourth_file_path)
            # Check if the required columns are in the fourth file
            if 'Last Ship CPN' in df_fourth.columns and 'Last Ship Date' in df_fourth.columns:
                # Group the data by 'Last Ship CPN' and get the latest (max) 'Last Ship Date' for each CPN
                latest_ship_dates = df_fourth.groupby('Last Ship CPN')['Last Ship Date'].max()
                # Update 'Last Ship Date' in the first DataFrame with the latest ship date where it is missing
                df_first['Last Ship Date'] = df_first.apply(
                    lambda row: row['Last Ship Date'] if pd.notna(row['Last Ship Date']) else latest_ship_dates.get(
                        row['CPN']), axis=1)
                return True
            else:
                # Display an error message if the required columns are not found
                messagebox.showerror("Error", "Required columns not found in the fourth file.")
                return False
        else:
            # Display a warning message if the file selection is cancelled
            messagebox.showwarning("Cancelled", "Fourth file open cancelled.")
            return False
    def save_file(df_first):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_file_path:
            # Save DataFrame to an Excel file
            df_first.to_excel(save_file_path, index=False)

            # Open the saved Excel file and add formatting
            workbook = load_workbook(save_file_path)
            worksheet = workbook.active

            # Apply wrap text for specific columns just going to do A for now and implement more later
            for row in worksheet.iter_rows(min_row=2):  # Skip the header row, currently not merging to control the columns being added 
                for cell in row:
                    if cell.column_letter in ['A']:
                        cell.alignment = Alignment(
                            wrap_text=False)  # Right now we are set to false considering it's not working properly

            # Enable filters for all columns
            worksheet.auto_filter.ref = worksheet.dimensions

            # Save the changes
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "Files processed, formatted, and saved successfully!")
        else:
            messagebox.showwarning("Cancelled", "File save cancelled.")

    # Add content specific to the Benchmark module
    label = tk.Label(plexus_window, text="Welcome Partnership Member", font=("Verdana", 24))
    label.pack(pady=20)

    # Add content specific to the Benchmark module
    my_font = font.Font(family="Verdana", size=20, underline=True)  # Define a font with underline
    label = tk.Label(plexus_window, text="Instructions:", font=my_font)
    label.pack(pady=20)

    # Add content specific to the Benchmark module
    label = tk.Label(plexus_window, text="First you will select the file where we have our CPN's\n"
                                         "Next we will select the Latest Contract File for Plexus\n"
                                         "Next we will select our Plexus Backlog File\n"
                                         "Lastly we will select out Plexus Sales History File\n"
                                         "Finally SAVE your final file", font=("Verdana", 20))
    label.pack(pady=20)

    # Button to open the file dialog
    open_file_btn = tk.Button(plexus_window, text="Select Excel Files", command=open_and_process_file)
    open_file_btn.pack(pady=10)
